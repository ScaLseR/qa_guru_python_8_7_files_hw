"""Домашняя работа: Создать новые тесты, которые заархивируют имеющиеся в resources различные типы файлов
(xls, xlsx, pdf, txt) в один архив, отправят его в tmp и проверят тестом в архиве каждый из файлов,
что он является тем самым, который был заархивирован, не распаковывая архив."""
import os
from utils import RESOURCES_PATH, TMP_PATH
from zipfile import ZipFile
from xlrd import open_workbook
import pytest
from openpyxl import *
from pypdf import PdfReader


def test_zip_files_from_resources_names():
    """Проверяем наличие в созданном архиве всех файлов по списку имен"""
    # получаем список имен файлов в resources
    file_in_dir = os.listdir(RESOURCES_PATH)
    with ZipFile(os.path.join(TMP_PATH, 'test.zip'), mode='r') as zf:
        assert file_in_dir == zf.namelist()


@pytest.mark.parametrize("file_name", ['Hello.txt'])
def test_zip_file_txt(file_name):
    """Проверяем соответствие размеров исходного txt и файла в архиве,
    + проверка файла по содержимому"""
    # получаем размер исходного файла
    txt_file_size = os.path.getsize(os.path.join(RESOURCES_PATH, file_name))
    # получаем содержимое исходного файла
    with open(os.path.join(RESOURCES_PATH, file_name), 'r') as f:
        txt_file_text = f.read()

    with ZipFile(os.path.join(TMP_PATH, 'test.zip'), mode='r') as zf:
        # проверяем на соответствие размеру файла
        assert txt_file_size == zf.getinfo(file_name).file_size
        # проверяем по содержимому
        assert txt_file_text in zf.read(file_name).decode()


@pytest.mark.parametrize("file_name", ['file_example_XLS_10.xls'])
def test_zip_file_xls(file_name):
    """Проверяем соответствие размеров исходного xls и файла в архиве,
    + проверка файла по содержимому"""
    # получаем размер исходного файла
    xls_file_size = os.path.getsize(os.path.join(RESOURCES_PATH, file_name))
    # получаем содержимое исходного xls файла
    book = open_workbook(os.path.join(RESOURCES_PATH, file_name))
    sheets_count = book.nsheets
    sheets_names = book.sheet_names()
    xls_file_text = book.sheet_by_index(0).cell_value(9, 3)

    with ZipFile(os.path.join(TMP_PATH, 'test.zip'), mode='r') as zf:
        # проверяем на соответствие размеру файла
        assert xls_file_size == zf.getinfo(file_name).file_size
        # получаем содержимое xls файла в zip архиве
        book_zip = open_workbook(file_contents=zf.read(file_name))
        # проверяем по количеству листов
        assert sheets_count == book_zip.nsheets
        # проверяем по названию листов
        assert sheets_names == book_zip.sheet_names()
        # проверяем по содержимому
        assert xls_file_text == book_zip.sheet_by_index(0).cell_value(9, 3)


@pytest.mark.parametrize("file_name", ['file_example_XLSX_50.xlsx'])
def test_zip_file_xlsx(file_name):
    """Проверяем соответствие размеров исходного xlsx и файла в архиве,
    + проверка файла по содержимому"""
    # получаем размер исходного файла
    xlsx_file_size = os.path.getsize(os.path.join(RESOURCES_PATH, file_name))
    # получаем содержимое исходного xlsx файла
    book = load_workbook(os.path.join(RESOURCES_PATH, file_name))
    sheets_count = len(book.sheetnames)
    sheets_names = book.sheetnames
    sheet = book.active
    xlsx_file_text = sheet.cell(9, 3).value

    with ZipFile(os.path.join(TMP_PATH, 'test.zip'), mode='r') as zf:
        # проверяем на соответствие размеру файла
        assert xlsx_file_size == zf.getinfo(file_name).file_size
        # получаем содержимое xlsx файла в zip архиве
        book_zip = load_workbook(zf.open(file_name, 'r'))
        # проверяем по количеству листов
        assert sheets_count == len(book_zip.sheetnames)
        # проверяем по названию листов
        assert sheets_names == book_zip.sheetnames
        # проверяем по содержимому
        sheet_zip = book_zip.active
        assert xlsx_file_text == sheet_zip.cell(9, 3).value


@pytest.mark.parametrize("file_name", ['Python Testing with Pytest (Brian Okken).pdf'])
def test_zip_file_pdf(file_name):
    """Проверяем соответствие размеров исходного pdf и файла в архиве,
    + проверка файла по содержимому"""
    # получаем размер исходного файла
    pdf_file_size = os.path.getsize(os.path.join(RESOURCES_PATH, file_name))
    # получаем содержимое исходного pdf файла
    reader = PdfReader(os.path.join(RESOURCES_PATH, file_name))
    pdf_page_count = len(reader.pages)
    pdf_page_text = reader.pages[1].extract_text()

    with ZipFile(os.path.join(TMP_PATH, 'test.zip'), mode='r') as zf:
        # проверяем на соответствие размеру файла
        assert pdf_file_size == zf.getinfo(file_name).file_size
        # получаем содержимое pdf файла в zip архиве
        zip_reader = PdfReader(zf.open(file_name, 'r'))
        # проверяем по количеству листов
        assert pdf_page_count == len(zip_reader.pages)
        # проверяем по содержимому
        assert pdf_page_text == zip_reader.pages[1].extract_text()
